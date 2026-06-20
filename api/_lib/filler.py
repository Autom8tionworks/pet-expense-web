"""Fill the real fixed master templates from a submission data model.

Design choices that matter:
* We NEVER call openpyxl.insert_rows() (see rows.py). Overflow is handled by
  expand_rows(), which copies styles and re-declares merges explicitly.
* After laying out rows we REGENERATE every formula at its final coordinate
  rather than trying to translate shifted references. This is what makes dynamic
  insertion safe: totals, per-diem INDEX/MATCH, mileage amounts and the
  section/amount-due roll-ups are all rewritten from the known layout.
* We do NOT reintroduce XLOOKUP -- per-diem meals use IFERROR(INDEX/MATCH),
  matching the already-fixed template.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation

from .model import JobSubmission, OfficeSubmission, ReportType
from .rows import expand_rows


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _clear_band(ws, r0: int, r1: int, c0: int, c1: int) -> None:
    """Blank out cell *values* (styles/merges untouched) across a rectangle."""
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):   # merged sub-cells are read-only
                cell.value = None


def _col(letter: str) -> int:
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def _sig_text(name: str, when: datetime) -> str:
    return f"{name}  (e-signed {when:%Y-%m-%d %H:%M})"


@dataclass
class FillResult:
    """Where the file landed + the final coordinates of key roll-up cells.
    Coordinates are returned because dynamic row insertion shifts them."""
    path: str
    sheet: str
    cells: dict          # logical name -> A1 coordinate (post-expansion)


# --------------------------------------------------------------------------- #
# JOB report  ->  2026_DomTravelER.xlsx
# --------------------------------------------------------------------------- #
def fill_job_report(sub: JobSubmission, template_path: str, out_path: str,
                    generated_at: datetime | None = None) -> str:
    errs = sub.validate()
    if errs:
        raise ValueError("Invalid job submission: " + "; ".join(errs))
    when = generated_at or datetime.now()

    wb = load_workbook(template_path)
    ws = wb["Dom. travel Exp Rpt"]

    # ---- original layout anchors (1-based rows) ----
    SEC1_FIRST = 10
    SEC1_CAP = 2
    COMMUTE_ROW = 12          # original
    RATE_ROW = 14             # original; holds B{rate} and Section-1 totals
    SEC2_FIRST = 17
    SEC2_CAP = 10
    SEC2_TOT = 27             # original totals row
    GRAND_ROW = 28            # original; J = grand total of section 2
    TOT_S1, TOT_S2, TOT_S3 = 30, 31, 32   # F column section roll-ups (original)
    AMT_DUE = 34              # original; F = amount due
    SIG_ROW = 34             # original; G = employee signature line (right block)
    SEC3_FIRST = 38
    SEC3_CAP = 2
    SEC3_TOT = 40            # original; J = misc total

    # ---- how many extra rows each section needs ----
    n1 = max(0, len(sub.mileage) - SEC1_CAP)
    n2 = max(0, len(sub.expenses) - SEC2_CAP)
    n3 = max(0, len(sub.misc_expenses) - SEC3_CAP)

    # ---- apply expansions top-to-bottom in CURRENT coordinates ----
    # Section 1: insert before the commute row, clone data-row 10's styling/merges
    if n1:
        expand_rows(ws, insert_at=COMMUTE_ROW, n=n1, template_row=SEC1_FIRST)
    # Section 2: totals row has shifted by n1
    if n2:
        expand_rows(ws, insert_at=SEC2_TOT + n1, n=n2, template_row=SEC2_FIRST + n1)
    # Section 3: shifted by n1 + n2
    if n3:
        expand_rows(ws, insert_at=SEC3_TOT + n1 + n2, n=n3,
                    template_row=SEC3_FIRST + n1 + n2)

    # ---- final coordinate mapping ----
    def f(r: int) -> int:
        return (r
                + (n1 if r >= COMMUTE_ROW else 0)
                + (n2 if r >= SEC2_TOT else 0)
                + (n3 if r >= SEC3_TOT else 0))

    sec1_first = SEC1_FIRST
    sec1_last = SEC1_FIRST + max(SEC1_CAP, len(sub.mileage)) - 1
    commute_row = f(COMMUTE_ROW)
    rate_row = f(RATE_ROW)
    sec2_first = f(SEC2_FIRST)
    sec2_last = sec2_first + max(SEC2_CAP, len(sub.expenses)) - 1
    sec2_tot = f(SEC2_TOT)
    grand_row = f(GRAND_ROW)
    sec3_first = f(SEC3_FIRST)
    sec3_last = sec3_first + max(SEC3_CAP, len(sub.misc_expenses)) - 1
    sec3_tot = f(SEC3_TOT)

    rate_ref = f"$B${rate_row}"

    # ---- header fields ----
    ws["B3"] = sub.submitter_name
    ws["G3"] = sub.customer_name_state
    ws["B4"] = sub.customer_billing_address
    ws["B5"] = sub.customer_po
    ws["H5"] = sub.customer_contact
    ws["B6"] = sub.job_number
    ws["G6"] = sub.overnight_travel
    if sub.days_traveled is not None:
        ws["J6"] = sub.days_traveled
    ws["B7"] = sub.project_description

    # ---- Section 1: personal-car mileage ----
    _clear_band(ws, sec1_first, sec1_last, _col("A"), _col("J"))
    for i, m in enumerate(sub.mileage):
        r = sec1_first + i
        ws.cell(r, _col("A")).value = m.travel_date
        ws.cell(r, _col("C")).value = m.origin
        ws.cell(r, _col("F")).value = m.destination
        ws.cell(r, _col("I")).value = m.miles
        ws.cell(r, _col("J")).value = f"=I{r}*{rate_ref}"
    # rows beyond the data but within original capacity keep the amount formula
    for r in range(sec1_first + len(sub.mileage), sec1_last + 1):
        ws.cell(r, _col("J")).value = f"=I{r}*{rate_ref}"
    # commute offset row
    co = sub.commute_offset
    ws.cell(commute_row, _col("E")).value = co.rt_commute_miles or None
    ws.cell(commute_row, _col("H")).value = co.num_days or None
    ws.cell(commute_row, _col("I")).value = f"=+E{commute_row}*H{commute_row}"
    ws.cell(commute_row, _col("J")).value = f"=-I{commute_row}*{rate_ref}"
    # section-1 totals
    ws.cell(rate_row, _col("I")).value = f"=SUM(I{sec1_first}:I{sec1_last})-I{commute_row}"
    ws.cell(rate_row, _col("J")).value = f"=+SUM(J{sec1_first}:J{commute_row})"

    # ---- Section 2: travel expenses ----
    _clear_band(ws, sec2_first, sec2_last, _col("A"), _col("J"))
    for i, e in enumerate(sub.expenses):
        r = sec2_first + i
        ws.cell(r, _col("A")).value = e.line_date
        ws.cell(r, _col("B")).value = e.per_diem.value
        ws.cell(r, _col("D")).value = e.airfare or None
        ws.cell(r, _col("E")).value = e.ground_transport or None
        ws.cell(r, _col("F")).value = e.car_rental_gas or None
        ws.cell(r, _col("G")).value = e.lodging or None
        ws.cell(r, _col("H")).value = e.telephone_fax or None
        ws.cell(r, _col("I")).value = e.parking or None
        ws.cell(r, _col("J")).value = e.misc or None
    # meals = per-diem lookup, for EVERY data row (IFERROR(INDEX/MATCH); NOT XLOOKUP)
    for r in range(sec2_first, sec2_last + 1):
        ws.cell(r, _col("C")).value = (
            f"=IFERROR(INDEX(per_diem_rates,MATCH(B{r},per_diem_type,0)),0)")
    # section-2 totals row
    for cl in "CDEFGHIJ":
        ws.cell(sec2_tot, _col(cl)).value = f"=SUM({cl}{sec2_first}:{cl}{sec2_last})"
    # grand total
    ws.cell(grand_row, _col("J")).value = f"=SUM(C{sec2_tot}:J{sec2_tot})"

    # extend the per-diem dropdown validation across the (possibly grown) band
    new_sqref = f"B{sec2_first}:B{sec2_last}"
    found = False
    for dv in list(ws.data_validations.dataValidation):
        if dv.type == "list":
            dv.sqref = new_sqref
            found = True
    if not found:
        dv = DataValidation(type="list", formula1="per_diem_type", allow_blank=True)
        dv.sqref = new_sqref
        ws.add_data_validation(dv)

    # ---- Section 3: miscellaneous expenses ----
    _clear_band(ws, sec3_first, sec3_last, _col("A"), _col("J"))
    for i, mx in enumerate(sub.misc_expenses):
        r = sec3_first + i
        ws.cell(r, _col("A")).value = mx.expense_date
        ws.cell(r, _col("C")).value = mx.description
        ws.cell(r, _col("J")).value = mx.amount or None
    ws.cell(sec3_tot, _col("J")).value = f"=SUM(J{sec3_first}:J{sec3_last})"

    # ---- roll-ups (Total Section 1/2/3 + Amount Due = 1+2+3) ----
    ws.cell(f(TOT_S1), _col("F")).value = f"=J{rate_row}"
    ws.cell(f(TOT_S2), _col("F")).value = f"=J{grand_row}"
    ws.cell(f(TOT_S3), _col("F")).value = f"=J{sec3_tot}"
    ws.cell(f(AMT_DUE), _col("F")).value = f"=SUM(F{f(TOT_S1)}:F{f(TOT_S3)})"

    # ---- electronic signature (no approval step) ----
    ws.cell(f(SIG_ROW), _col("G")).value = _sig_text(sub.submitter_name, when)

    wb.save(out_path)
    return FillResult(out_path, ws.title, {
        "total_s1": f"F{f(TOT_S1)}", "total_s2": f"F{f(TOT_S2)}",
        "total_s3": f"F{f(TOT_S3)}", "amount_due": f"F{f(AMT_DUE)}",
        "grand_total": f"J{grand_row}", "sec1_total": f"J{rate_row}",
    })


# --------------------------------------------------------------------------- #
# OFFICE report  ->  2026_OER.xlsx
# --------------------------------------------------------------------------- #
def fill_office_report(sub: OfficeSubmission, template_path: str, out_path: str,
                       generated_at: datetime | None = None) -> str:
    errs = sub.validate()
    if errs:
        raise ValueError("Invalid office submission: " + "; ".join(errs))
    when = generated_at or datetime.now()

    wb = load_workbook(template_path)
    ws = wb["Office Exp Rpt"]

    # ---- original anchors ----
    SEC1_FIRST = 6
    SEC1_CAP = 4
    RATE_TOT_ROW = 10        # B = rate, H/I = section-1 totals
    SEC2_FIRST = 13
    SEC2_CAP = 11
    SEC2_TOT = 24            # I = grand total
    SEC3_FIRST = 27
    SEC3_CAP = 3
    SEC3_TOT = 30            # I = grand total
    TOT_S1, TOT_S2, TOT_S3 = 32, 33, 34   # E column (original)
    AMT_DUE = 36             # E (original)
    SIG_ROW = 35            # F = employee signature line (original)

    n1 = max(0, len(sub.mileage) - SEC1_CAP)
    n2 = max(0, len(sub.office_expenses) - SEC2_CAP)
    n3 = max(0, len(sub.credit_card_charges) - SEC3_CAP)

    if n1:
        expand_rows(ws, insert_at=RATE_TOT_ROW, n=n1, template_row=SEC1_FIRST)
    if n2:
        expand_rows(ws, insert_at=SEC2_TOT + n1, n=n2, template_row=SEC2_FIRST + n1)
    if n3:
        expand_rows(ws, insert_at=SEC3_TOT + n1 + n2, n=n3,
                    template_row=SEC3_FIRST + n1 + n2)

    def f(r: int) -> int:
        return (r
                + (n1 if r >= RATE_TOT_ROW else 0)
                + (n2 if r >= SEC2_TOT else 0)
                + (n3 if r >= SEC3_TOT else 0))

    sec1_first = SEC1_FIRST
    sec1_last = SEC1_FIRST + max(SEC1_CAP, len(sub.mileage)) - 1
    rate_row = f(RATE_TOT_ROW)
    sec2_first = f(SEC2_FIRST)
    sec2_last = sec2_first + max(SEC2_CAP, len(sub.office_expenses)) - 1
    sec2_tot = f(SEC2_TOT)
    sec3_first = f(SEC3_FIRST)
    sec3_last = sec3_first + max(SEC3_CAP, len(sub.credit_card_charges)) - 1
    sec3_tot = f(SEC3_TOT)
    rate_ref = f"$B${rate_row}"

    # ---- header ----
    ws["B3"] = sub.submitter_name
    if sub.report_date is not None:
        ws["F3"] = f"Date:  {sub.report_date:%B %-d, %Y}"

    # ---- Section 1 mileage ----
    _clear_band(ws, sec1_first, sec1_last, _col("A"), _col("I"))
    for i, m in enumerate(sub.mileage):
        r = sec1_first + i
        ws.cell(r, _col("A")).value = m.travel_date
        ws.cell(r, _col("B")).value = m.origin
        ws.cell(r, _col("E")).value = m.destination
        ws.cell(r, _col("H")).value = m.miles
        ws.cell(r, _col("I")).value = f"=H{r}*{rate_ref}"
    for r in range(sec1_first + len(sub.mileage), sec1_last + 1):
        ws.cell(r, _col("I")).value = f"=H{r}*{rate_ref}"
    ws.cell(rate_row, _col("H")).value = f"=SUM(H{sec1_first}:H{sec1_last})"
    ws.cell(rate_row, _col("I")).value = f"=SUM(I{sec1_first}:I{sec1_last})"

    # ---- Section 2 office expenses ----
    _clear_band(ws, sec2_first, sec2_last, _col("A"), _col("I"))
    for i, oe in enumerate(sub.office_expenses):
        r = sec2_first + i
        ws.cell(r, _col("A")).value = oe.expense_date
        ws.cell(r, _col("B")).value = oe.description
        ws.cell(r, _col("I")).value = oe.cost or None
    ws.cell(sec2_tot, _col("I")).value = f"=SUM(I{sec2_first}:I{sec2_last})"

    # ---- Section 3 credit-card charges ----
    _clear_band(ws, sec3_first, sec3_last, _col("A"), _col("I"))
    for i, cc in enumerate(sub.credit_card_charges):
        r = sec3_first + i
        ws.cell(r, _col("A")).value = cc.charge_date
        ws.cell(r, _col("B")).value = cc.vendor
        ws.cell(r, _col("F")).value = cc.description
        ws.cell(r, _col("I")).value = cc.amount or None
    ws.cell(sec3_tot, _col("I")).value = f"=SUM(I{sec3_first}:I{sec3_last})"

    # ---- roll-ups ----
    ws.cell(f(TOT_S1), _col("E")).value = f"=I{rate_row}"
    ws.cell(f(TOT_S2), _col("E")).value = f"=I{sec2_tot}"
    ws.cell(f(TOT_S3), _col("E")).value = f"=I{sec3_tot}"
    ws.cell(f(AMT_DUE), _col("E")).value = f"=SUM(E{f(TOT_S1)}:E{f(TOT_S3)})"

    # ---- electronic signature ----
    ws.cell(f(SIG_ROW), _col("F")).value = _sig_text(sub.submitter_name, when)

    wb.save(out_path)
    return FillResult(out_path, ws.title, {
        "total_s1": f"E{f(TOT_S1)}", "total_s2": f"E{f(TOT_S2)}",
        "total_s3": f"E{f(TOT_S3)}", "amount_due": f"E{f(AMT_DUE)}",
    })


# --------------------------------------------------------------------------- #
# dispatcher
# --------------------------------------------------------------------------- #
def fill_report(sub, template_path: str, out_path: str, generated_at=None) -> str:
    if isinstance(sub, JobSubmission) or getattr(sub, "report_type", None) == ReportType.JOB:
        return fill_job_report(sub, template_path, out_path, generated_at)
    return fill_office_report(sub, template_path, out_path, generated_at)
