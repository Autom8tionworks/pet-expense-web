"""Safe dynamic row insertion for heavily-merged templates.

WHY THIS EXISTS (discovered gotcha):
openpyxl's Worksheet.insert_rows() silently drops cell values adjacent to merged
ranges and does NOT move merged-cell ranges along with the shifted content. On
these expense templates -- which are dense with per-row merges -- that corrupts
the sheet. The safe pattern is to do the shift ourselves: copy each cell's value
and full style to its new position, then explicitly re-declare every merged range
at its shifted location. Per-row merges in the data band are re-created for the
newly opened rows by cloning a representative template data row.

This module exposes ONE primitive, `expand_rows`, used by the filler.
"""
from __future__ import annotations

from copy import copy
from openpyxl.worksheet.worksheet import Worksheet


def _row_merge_colspans(ws: Worksheet, row: int) -> list[tuple[int, int]]:
    """Column spans of merges that live entirely on a single `row`."""
    spans = []
    for m in ws.merged_cells.ranges:
        if m.min_row == row and m.max_row == row:
            spans.append((m.min_col, m.max_col))
    return spans


def expand_rows(ws: Worksheet, insert_at: int, n: int, template_row: int) -> None:
    """Insert `n` blank, correctly-styled rows starting at `insert_at`.

    Everything originally at row >= insert_at moves down by `n`. The opened gap
    rows [insert_at, insert_at+n-1] inherit the cell styles, row height, and
    per-row merge col-spans of `template_row` (a representative existing data
    row). All other merges are preserved and shifted where appropriate.

    Coordinates are CURRENT worksheet coordinates at call time. `template_row`
    must be a row that is NOT inside the gap being opened.
    """
    if n <= 0:
        return

    max_row = ws.max_row
    max_col = ws.max_column

    # 1) Snapshot the template data row's styles + per-row merge col-spans BEFORE
    #    we mutate anything. (We copy style objects so later shifts can't alias.)
    tmpl_styles = {c: copy(ws.cell(row=template_row, column=c)._style)
                   for c in range(1, max_col + 1)}
    tmpl_merge_spans = _row_merge_colspans(ws, template_row)
    tmpl_height = ws.row_dimensions[template_row].height

    # 2) Record every merge, then clear them all (we re-apply explicitly).
    original_merges = [(m.min_row, m.min_col, m.max_row, m.max_col)
                       for m in ws.merged_cells.ranges]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(start_row=m.min_row, start_column=m.min_col,
                         end_row=m.max_row, end_column=m.max_col)

    # 3) Move rows down, bottom-to-top, copying value + full style.
    for r in range(max_row, insert_at - 1, -1):
        src_h = ws.row_dimensions[r].height
        for c in range(1, max_col + 1):
            src = ws.cell(row=r, column=c)
            dst = ws.cell(row=r + n, column=c)
            dst.value = src.value
            if src.has_style:
                dst._style = copy(src._style)
        if src_h is not None:
            ws.row_dimensions[r + n].height = src_h

    # 4) Clear the opened gap rows and apply template-row styling.
    for gap in range(insert_at, insert_at + n):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=gap, column=c)
            cell.value = None
            cell._style = copy(tmpl_styles[c])
        if tmpl_height is not None:
            ws.row_dimensions[gap].height = tmpl_height

    # 5) Re-apply merges. Any merge starting at/after insert_at shifts by n.
    for (r0, c0, r1, c1) in original_merges:
        if r0 >= insert_at:
            r0 += n
            r1 += n
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)

    # 6) Clone the template row's per-row merges onto each new gap row.
    for gap in range(insert_at, insert_at + n):
        for (c0, c1) in tmpl_merge_spans:
            if c1 > c0:
                ws.merge_cells(start_row=gap, start_column=c0,
                               end_row=gap, end_column=c1)
